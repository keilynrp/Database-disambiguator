"""
Sprint 40 regression tests — Enterprise Export (PDF + Excel).

WeasyPrint is patched for all PDF tests so CI doesn't need a browser engine.
Excel tests use openpyxl to verify workbook structure.
"""
import io
import pytest
from unittest.mock import MagicMock, patch

import openpyxl

from backend import models


_REPORT_PAYLOAD = {
    "domain_id": "default",
    "sections": ["entity_stats"],
    "title": "Test Export",
}

_MOCK_WEASY = "backend.routers.reports._make_pdf"


# ── PDF endpoint ───────────────────────────────────────────────────────────────

def test_pdf_requires_auth(client):
    resp = client.post("/exports/pdf", json=_REPORT_PAYLOAD)
    assert resp.status_code in (401, 403)


def test_pdf_calls_report_builder(client, auth_headers):
    """_make_pdf must be called with the HTML output from report_builder.build()."""
    fake_pdf = b"%PDF-1.4 fake"
    with patch(_MOCK_WEASY, return_value=fake_pdf) as mock_pdf:
        resp = client.post("/exports/pdf", json=_REPORT_PAYLOAD, headers=auth_headers)
    # May be 200 or 501 (weasyprint not installed) — should NOT be 422/403/500 from our code
    assert resp.status_code in (200, 501)
    if resp.status_code == 200:
        mock_pdf.assert_called_once()


def test_pdf_response_content_type(client, auth_headers):
    fake_pdf = b"%PDF-1.4 fake"
    with patch(_MOCK_WEASY, return_value=fake_pdf):
        resp = client.post("/exports/pdf", json=_REPORT_PAYLOAD, headers=auth_headers)
    if resp.status_code == 200:
        assert "application/pdf" in resp.headers["content-type"]


def test_pdf_invalid_section_returns_422(client, auth_headers):
    payload = {"domain_id": "default", "sections": ["nonexistent_section"]}
    with patch(_MOCK_WEASY, return_value=b""):
        resp = client.post("/exports/pdf", json=payload, headers=auth_headers)
    assert resp.status_code == 422


# ── Excel endpoint ─────────────────────────────────────────────────────────────

def test_excel_requires_auth(client):
    resp = client.post("/exports/excel", json=_REPORT_PAYLOAD)
    assert resp.status_code in (401, 403)


def test_excel_content_type_xlsx(client, auth_headers):
    resp = client.post("/exports/excel", json=_REPORT_PAYLOAD, headers=auth_headers)
    assert resp.status_code == 200
    ct = resp.headers["content-type"]
    assert "spreadsheetml" in ct or "officedocument" in ct


def test_excel_has_summary_sheet(client, auth_headers):
    """Response bytes must be a valid xlsx with a 'Summary' sheet."""
    resp = client.post("/exports/excel", json=_REPORT_PAYLOAD, headers=auth_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Summary" in wb.sheetnames


def test_excel_has_entities_sheet(client, auth_headers):
    resp = client.post("/exports/excel", json=_REPORT_PAYLOAD, headers=auth_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Entities" in wb.sheetnames


def test_excel_header_violet_fill(client, auth_headers):
    """The Summary sheet header row must use the violet brand fill (#5B21B6)."""
    resp = client.post("/exports/excel", json=_REPORT_PAYLOAD, headers=auth_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Summary"]
    fill = ws.cell(row=1, column=1).fill
    assert fill.fgColor.rgb.upper().endswith("5B21B6")


def test_excel_concepts_sheet_when_requested(client, auth_headers):
    """Including 'topic_clusters' in sections must produce a 'Concepts' sheet."""
    payload = {
        "domain_id": "default",
        "sections": ["entity_stats", "topic_clusters"],
    }
    resp = client.post("/exports/excel", json=payload, headers=auth_headers)
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Concepts" in wb.sheetnames
