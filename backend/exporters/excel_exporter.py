"""
Enterprise Excel exporter — multi-sheet branded workbook.
Used by POST /exports/excel in backend/routers/reports.py.
"""
from __future__ import annotations

import logging
from io import BytesIO
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from backend import models
from backend.analyzers.topic_modeling import TopicAnalyzer

logger = logging.getLogger(__name__)

# Violet brand palette (hex, no #)
_HEADER_FG   = "5B21B6"   # violet-800
_HEADER_FONT = "FFFFFF"   # white

_HEADER_FILL = PatternFill("solid", fgColor=_HEADER_FG)
_HEADER_FONT_STYLE = Font(color=_HEADER_FONT, bold=True, size=11)
_SUBHEADER_FONT = Font(bold=True, size=10)


def _style_header_row(ws, cols: list[str]) -> None:
    """Write and style a header row (row 1) with violet fill + white bold text."""
    for col_idx, header in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT_STYLE
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autofit(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Approximate column width based on content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


class EnterpriseExcelExporter:
    """Build a branded multi-sheet Excel workbook and return raw bytes."""

    _ENTITY_CAP = 5_000
    _CONCEPT_CAP = 50

    def build(self, db: Session, domain_id: str, sections: List[str]) -> bytes:
        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary KPIs ──────────────────────────────────────────────
        ws_summary = wb.active
        assert ws_summary is not None
        ws_summary.title = "Summary"
        self._write_summary(ws_summary, db, domain_id)

        # ── Sheet 2: Entities ──────────────────────────────────────────────────
        ws_entities = wb.create_sheet("Entities")
        self._write_entities(ws_entities, db)

        # ── Sheet 3: Concepts ─────────────────────────────────────────────────
        if "topic_clusters" in sections:
            ws_concepts = wb.create_sheet("Concepts")
            self._write_concepts(ws_concepts, db, domain_id)

        # ── Sheet 4: Harmonization Log ────────────────────────────────────────
        if "harmonization_log" in sections:
            ws_harm = wb.create_sheet("Harmonization")
            self._write_harmonization(ws_harm, db)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Private sheet writers ──────────────────────────────────────────────────

    def _write_summary(self, ws, db: Session, domain_id: str) -> None:
        total = db.query(models.RawEntity).count()
        enriched = (
            db.query(models.RawEntity)
            .filter(models.RawEntity.enrichment_status == "completed")
            .count()
        )
        pct = round(enriched / total * 100, 1) if total > 0 else 0.0

        from sqlalchemy import func
        avg_row = (
            db.query(func.avg(models.RawEntity.enrichment_citation_count))
            .filter(models.RawEntity.enrichment_status == "completed")
            .scalar()
        )
        avg_cit = round(float(avg_row), 1) if avg_row is not None else 0.0

        headers = ["Metric", "Value"]
        _style_header_row(ws, headers)

        rows = [
            ("Active Domain",       domain_id),
            ("Total Entities",      total),
            ("Enriched Entities",   enriched),
            ("Enrichment %",        f"{pct}%"),
            ("Avg Citations",       avg_cit),
            ("Platform",            "UKIP — Universal Knowledge Intelligence Platform"),
        ]
        for row_idx, (metric, value) in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=metric).font = _SUBHEADER_FONT
            ws.cell(row=row_idx, column=2, value=value)

        _autofit(ws)

    def _write_entities(self, ws, db: Session) -> None:
        headers = [
            "ID", "Primary Label", "Secondary Label", "Canonical ID", "Entity Type",
            "Enrichment Status", "Citation Count", "Source",
        ]
        _style_header_row(ws, headers)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        rows = (
            db.query(models.RawEntity)
            .order_by(models.RawEntity.id)
            .limit(self._ENTITY_CAP)
            .all()
        )
        for row_idx, e in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=e.id)
            ws.cell(row=row_idx, column=2, value=e.primary_label)
            ws.cell(row=row_idx, column=3, value=e.secondary_label)
            ws.cell(row=row_idx, column=4, value=e.canonical_id)
            ws.cell(row=row_idx, column=5, value=e.entity_type)
            ws.cell(row=row_idx, column=6, value=e.enrichment_status)
            ws.cell(row=row_idx, column=7, value=e.enrichment_citation_count)
            ws.cell(row=row_idx, column=8, value=e.enrichment_source)

        _autofit(ws)

    def _write_concepts(self, ws, db: Session, domain_id: str) -> None:  # db kept for consistency
        headers = ["Rank", "Concept", "Count", "Percentage (%)"]
        _style_header_row(ws, headers)

        try:
            result = TopicAnalyzer().top_topics(domain_id=domain_id, top_n=self._CONCEPT_CAP)
            topics = result.get("topics", [])
        except Exception:
            logger.warning("TopicAnalyzer failed in excel_exporter", exc_info=True)
            topics = []

        for row_idx, t in enumerate(topics, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=t.get("concept", ""))
            ws.cell(row=row_idx, column=3, value=t.get("count", 0))
            ws.cell(row=row_idx, column=4, value=round(t.get("pct", 0.0), 2))

        _autofit(ws)

    def _write_harmonization(self, ws, db: Session) -> None:
        headers = ["ID", "Step ID", "Step Name", "Records Updated", "Fields Modified", "Executed At", "Reverted"]
        _style_header_row(ws, headers)

        rows = (
            db.query(models.HarmonizationLog)
            .order_by(models.HarmonizationLog.id.desc())
            .limit(200)
            .all()
        )
        for row_idx, h in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=h.id)
            ws.cell(row=row_idx, column=2, value=h.step_id)
            ws.cell(row=row_idx, column=3, value=h.step_name)
            ws.cell(row=row_idx, column=4, value=h.records_updated)
            ws.cell(row=row_idx, column=5, value=h.fields_modified)
            ws.cell(row=row_idx, column=6, value=str(h.executed_at) if h.executed_at else "")
            ws.cell(row=row_idx, column=7, value="Yes" if h.reverted else "No")

        _autofit(ws)
